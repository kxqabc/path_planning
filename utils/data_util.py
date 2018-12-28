#!/usr/bin/python
# -*- coding: utf-8 -*-

import os

import numpy as np
from openpyxl import Workbook

TSP_TAG = ('NODE_COORD_SECTION', 'TOUR_SECTION')


def joint(path, file_name):
    if path.endswith(os.sep):
        file_name = path + file_name
    else:
        file_name = os.sep.join((path, file_name))
    return file_name


def get_date_from_tsp(path, file_name, data_dim):
    file_name = joint(path, file_name)
    info = dict()
    points = []
    with open(file_name, 'r') as f:
        lines = f.readlines()
        for line in lines:
            line = line.strip()     # 去除两端的空格
            if line.find(":") != -1:
                info_tag, info_content = line.split(":")
                info[info_tag] = info_content
            elif line in TSP_TAG:
                pass
            elif line == 'EOF':
                break
            else:
                point = parse_data_line(line, data_dim)
                points.append(point)
    point_array = np.asarray(points, dtype=float)
    return info, point_array


def parse_data_line(line, dim):
    if not line or int(dim) <= 0:
        raise ValueError("data line is empty or the dim is invalid!")
    axes_list = line.split(" ")
    if len(axes_list) != dim:
        raise ValueError("data dim is error!")
    # 去除index，只保留数据
    if dim > 1:
        axes_list = axes_list[1:]
    return [float(axes) for axes in axes_list]


def get_opt_result(path, file_name, data_dim):
    return get_date_from_tsp(path, file_name, data_dim)


def save_excel(path, file_name, dict_list, sheet_name=None):
    file_name = joint(path, file_name)
    work_book = Workbook()
    # get sheet by name, get active sheet if sheet_name is None
    if sheet_name:
        if sheet_name in work_book.sheetnames:
            sheet = work_book[sheet_name]
        else:
            sheet = work_book.create_sheet(title=sheet_name)
    else:
        sheet = work_book.active
    # add data(dict format) to sheet
    for data_dict in dict_list:
        sheet.append(data_dict)
    # save
    work_book.save(file_name + '.xlsx')
